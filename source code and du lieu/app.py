import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.io as pio
import matplotlib.pyplot as plt
import mysql.connector
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
import base64
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Đăng ký font Times New Roman hỗ trợ tiếng Việt
try:
    pdfmetrics.registerFont(TTFont('TimesNewRoman', 'C:/Windows/Fonts/times.ttf'))
except Exception as e:
    st.error(f"Không thể đăng ký font Times New Roman: {e}. Sử dụng font mặc định.")
    pdfmetrics.registerFont(TTFont('Helvetica', 'Helvetica'))  # Font mặc định của reportlab

# --- Định nghĩa các Hằng số và Hàm Hỗ trợ ---

CRITERIA_MAP = {
    "phobien": "Phổ biến",
    "nhipdo": "Nhịp độ",
    "nangluong": "Năng lượng",
    "nhay": "Nhảy",
    "camxuc": "Cảm xúc"
}
CRITERIA_IDS = list(CRITERIA_MAP.keys())
CRITERIA_NAMES = list(CRITERIA_MAP.values())

RI_LOOKUP = {
    1: 0, 2: 0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32,
    8: 1.41, 9: 1.45, 10: 1.49
}

AHP_SCALE = {
    1: "1: Bằng nhau",
    2: "2",
    3: "3: Quan trọng hơn một chút",
    4: "4",
    5: "5: Quan trọng hơn",
    6: "6",
    7: "7: Rất quan trọng",
    8: "8",
    9: "9: Cực kỳ quan trọng",
    1/2: "1/2",
    1/3: "1/3: Kém quan trọng hơn một chút",
    1/4: "1/4",
    1/5: "1/5: Kém quan trọng hơn",
    1/6: "1/6",
    1/7: "1/7: Kém rất quan trọng",
    1/8: "1/8",
    1/9: "1/9: Cực kỳ kém quan trọng",
}
AHP_VALUES = list(AHP_SCALE.keys())
AHP_LABELS = [f"{AHP_SCALE[v]} ({v:.2f})" for v in AHP_VALUES]

def calculate_ahp_weights_consistency(comparison_matrix):
    n = comparison_matrix.shape[0]
    if n == 0:
        return np.array([]), 0, 0, 0
    column_sums = comparison_matrix.sum(axis=0)
    column_sums[column_sums == 0] = 1e-9
    normalized_matrix = comparison_matrix / column_sums
    weights = normalized_matrix.mean(axis=1)
    if n <= 1:
        return weights, n, 0, 0
    weighted_sum_vector = comparison_matrix @ weights
    consistency_vector = weighted_sum_vector / weights
    lambda_max = np.mean(consistency_vector)
    if n == 2:
        ci = 0
        cr = 0.0 if np.isclose(lambda_max, 2) else float('inf')
    else:
        ci = (lambda_max - n) / (n - 1)
        ri = RI_LOOKUP.get(n)
        if ri is None or ri == 0:
            cr = float('inf') if not np.isclose(ci, 0) else 0.0
        else:
            cr = ci / ri
    return weights, lambda_max, ci, cr

def create_criteria_matrix_from_inputs(inputs, criteria_ids):
    n = len(criteria_ids)
    matrix = np.ones((n, n))
    idx = 0
    for i in range(n):
        for j in range(i + 1, n):
            value = inputs[idx]
            matrix[i, j] = value
            matrix[j, i] = 1.0 / value if value != 0 else 1e9
            idx += 1
    return matrix

def create_alternative_comparison_matrix(values, criterion_type):
    n = len(values)
    matrix = np.ones((n, n), dtype=float)
    epsilon = 1e-9
    normalized_values = values.copy()
    
    # Chuẩn hóa giá trị theo loại tiêu chí
    if criterion_type == "phobien":  # Tiêu chí Phổ biến (track_popularity: 0-100)
        normalized_values = values / 100.0  # Chuẩn hóa về thang 0-1 giống energy
    elif criterion_type in ["nangluong", "nhay", "camxuc"]:  # Tiêu chí đã ở thang 0-1
        normalized_values = values
    elif criterion_type == "nhipdo":  # Tiêu chí Nhịp độ (tempo: giá trị dương)
        normalized_values = values / np.max(values) if np.max(values) > 0 else values
    
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            numerator = float(normalized_values[i]) if isinstance(normalized_values[i], (int, float)) and not np.isnan(normalized_values[i]) else 0.0
            denominator = float(normalized_values[j]) if isinstance(normalized_values[j], (int, float)) and not np.isnan(normalized_values[j]) else epsilon
            matrix[i, j] = numerator / denominator if denominator != 0 else epsilon
    return matrix

@st.cache_data
def load_data(filepath):
    try:
        # Đọc file CSV với các cột mong đợi
        df = pd.read_csv(filepath)
        
        # Danh sách cột bắt buộc
        required_columns = ['track_name', 'track_artist', 'track_popularity', 'danceability', 'energy', 'valence', 'tempo']
        
        # Kiểm tra nếu thiếu cột nào
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"Lỗi: File CSV thiếu các cột bắt buộc: {', '.join(missing_columns)}. "
                     f"Vui lòng đảm bảo file chứa tất cả các cột: {', '.join(required_columns)}.")
            return None

        # Kiểm tra kiểu dữ liệu và phạm vi giá trị
        errors = []
        df_selected = df[required_columns].copy()

        # Kiểm tra cột chuỗi (track_name, track_artist)
        for col in ['track_name', 'track_artist']:
            if df_selected[col].isnull().all() or df_selected[col].eq('').all():
                errors.append(f"Cột '{col}' không chứa dữ liệu hợp lệ (tất cả giá trị trống hoặc NULL).")

        # Kiểm tra cột số (numeric columns)
        numeric_cols = ['track_popularity', 'danceability', 'energy', 'valence', 'tempo']
        for col in numeric_cols:
            # Chuyển đổi sang kiểu số, lỗi sẽ thành NaN
            df_selected[col] = pd.to_numeric(df_selected[col], errors='coerce')
            # Kiểm tra giá trị NaN (dữ liệu không phải số)
            invalid_rows = df_selected[df_selected[col].isna()].index
            if not invalid_rows.empty:
                errors.append(f"Cột '{col}' chứa giá trị không phải số ở các dòng: {list(invalid_rows + 2)} "
                             f"(dòng 1 là tiêu đề). Ví dụ: '{df_selected.loc[invalid_rows[0], col]}'.")
            # Kiểm tra phạm vi giá trị
            if col == 'track_popularity':
                out_of_range = df_selected[(df_selected[col] < 0) | (df_selected[col] > 100)].index
                if not out_of_range.empty:
                    errors.append(f"Cột '{col}' phải từ 0 đến 100, nhưng có giá trị ngoài phạm vi ở các dòng: "
                                 f"{list(out_of_range + 2)}. Ví dụ: '{df_selected.loc[out_of_range[0], col]}'.")
            elif col in ['danceability', 'energy', 'valence']:
                out_of_range = df_selected[(df_selected[col] < 0) | (df_selected[col] > 1)].index
                if not out_of_range.empty:
                    errors.append(f"Cột '{col}' phải từ 0 đến 1, nhưng có giá trị ngoài phạm vi ở các dòng: "
                                 f"{list(out_of_range + 2)}. Ví dụ: '{df_selected.loc[out_of_range[0], col]}'.")
            elif col == 'tempo':
                out_of_range = df_selected[df_selected[col] < 0].index
                if not out_of_range.empty:
                    errors.append(f"Cột '{col}' phải lớn hơn hoặc bằng 0, nhưng có giá trị âm ở các dòng: "
                                 f"{list(out_of_range + 2)}. Ví dụ: '{df_selected.loc[out_of_range[0], col]}'.")

        # Nếu có lỗi, báo cho người dùng và dừng
        if errors:
            st.error("Lỗi: File CSV không đúng định dạng. Vui lòng kiểm tra các vấn đề sau:\n" + "\n".join(errors))
            st.warning("Định dạng mong đợi: File CSV phải chứa các cột 'track_name', 'track_artist', 'track_popularity' (0-100), "
                       "'danceability' (0-1), 'energy' (0-1), 'valence' (0-1), 'tempo' (>=0) với dữ liệu hợp lệ.")
            return None

        # Lọc dữ liệu không rỗng và chuẩn hóa
        df_selected.dropna(subset=['track_name', 'track_artist'], inplace=True)
        for col in numeric_cols:
            df_selected[col] = df_selected[col].fillna(0.0)

        return df_selected

    except FileNotFoundError:
        st.error(f"Lỗi: Không tìm thấy file '{filepath}'.")
        return None
    except Exception as e:
        st.error(f"Lỗi khi đọc file CSV: {e}")
        return None

def init_mysql_connection():
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Vinh115",
            database="song_ranking_db",
            port="3308"
        )
        return connection
    except mysql.connector.Error as e:
        st.error(f"Lỗi kết nối MySQL: {e}")
        return None

def create_results_table(connection):
    try:
        cursor = connection.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS song_ranking_results (
                id INT AUTO_INCREMENT PRIMARY KEY,
                `rank` INT NOT NULL,
                track_name VARCHAR(255) NOT NULL,
                track_artist VARCHAR(255) NOT NULL,
                final_score FLOAT NOT NULL,
                saved_at DATETIME NOT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        connection.commit()
        cursor.close()
    except mysql.connector.Error as e:
        st.error(f"Lỗi khi tạo bảng: {e}")

def save_results_to_mysql(connection, final_results_df):
    try:
        cursor = connection.cursor()
        saved_at = datetime.now()
        for _, row in final_results_df.iterrows():
            cursor.execute("""
                INSERT INTO song_ranking_results (`rank`, track_name, track_artist, final_score, saved_at)
                VALUES (%s, %s, %s, %s, %s)
            """, (int(row['Rank']), row['track_name'], row['track_artist'], float(row['Final Score']), saved_at))
        connection.commit()
        cursor.close()
        st.success("Dữ liệu đã được lưu thành công vào cơ sở dữ liệu MySQL!")
    except mysql.connector.Error as e:
        st.error(f"Lỗi khi lưu dữ liệu vào MySQL: {e}")

def load_history_from_mysql(connection):
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT `rank`, track_name, track_artist, final_score, saved_at
            FROM song_ranking_results
            ORDER BY saved_at DESC
        """)
        rows = cursor.fetchall()
        columns = ['Rank', 'track_name', 'track_artist', 'Final Score', 'saved_at']
        history_df = pd.DataFrame(rows, columns=columns)
        cursor.close()
        return history_df
    except mysql.connector.Error as e:
        st.error(f"Lỗi khi tải lịch sử từ MySQL: {e}")
        return pd.DataFrame()

def export_to_excel(criteria_matrix_df, criteria_weights_df, local_weights_dfs, final_results_df, figures):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "AHP Song Ranking Results"

    # Current row to keep track of where to write
    current_row = 1

    # Write User Inputs
    ws.cell(row=current_row, column=1, value="Đầu vào người dùng")
    current_row += 1
    ws.append(["Cặp Tiêu chí", "Giá trị"])
    idx = 0
    for i in range(len(CRITERIA_IDS)):
        for j in range(i + 1, len(CRITERIA_IDS)):
            crit_left = CRITERIA_MAP[CRITERIA_IDS[i]]
            crit_right = CRITERIA_MAP[CRITERIA_IDS[j]]
            ws.append([f"{crit_left} vs {crit_right}", st.session_state.get('pairwise_inputs', [])[idx]])
            idx += 1
    current_row += len(CRITERIA_IDS) * (len(CRITERIA_IDS) - 1) // 2 + 1

    # Add empty row
    current_row += 1

    # Write Criteria Comparison Matrix
    ws.cell(row=current_row, column=1, value="Ma trận So sánh Tiêu chí")
    current_row += 1
    ws.append([""] + CRITERIA_NAMES)
    for i, row in enumerate(criteria_matrix_df.values):
        formatted_row = []
        for x in row:
            if isinstance(x, (int, float)) and not np.isnan(x):
                formatted_row.append(f"{x:.3f}")
            else:
                formatted_row.append(str(x))
        ws.append([CRITERIA_NAMES[i]] + formatted_row)
    current_row += len(CRITERIA_NAMES) + 1

    # Add empty row
    current_row += 1

    # Write Criteria Weights
    ws.cell(row=current_row, column=1, value="Trọng số Tiêu chí")
    current_row += 1
    ws.append(["Tiêu chí", "Trọng số (Weights)", "Rank"])
    for _, row in criteria_weights_df.iterrows():
        weight = row['Trọng số (Weights)']
        formatted_weight = f"{weight:.4f}" if isinstance(weight, (int, float)) and not np.isnan(weight) else str(weight)
        ws.append([row['Tiêu chí'], formatted_weight, int(row['Rank'])])
    current_row += len(CRITERIA_NAMES) + 1

    # Add empty row
    current_row += 1

    # Write Alternative Comparison Matrices and Local Weights
    for crit_name, local_weights_df in local_weights_dfs.items():
        ws.cell(row=current_row, column=1, value=f"Ma trận So sánh Phương án - {crit_name}")
        current_row += 1
        ws.append([""] + local_weights_df['Song ID'].tolist())
        for i, row in enumerate(local_weights_df['Matrix'].values):
            formatted_row = []
            for x in row:
                if isinstance(x, (int, float)) and not np.isnan(x):
                    formatted_row.append(f"{x:.3f}")
                else:
                    formatted_row.append(str(x))
            ws.append([local_weights_df['Song ID'][i]] + formatted_row)
        current_row += len(local_weights_df) + 1
        ws.cell(row=current_row, column=1, value=f"Trọng số Phương án - {crit_name}")
        current_row += 1
        ws.append(["Song ID", "Weight"])
        for _, row in local_weights_df.iterrows():
            weight = row['Weight']
            formatted_weight = f"{weight:.4f}" if isinstance(weight, (int, float)) and not np.isnan(weight) else str(weight)
            ws.append([row['Song ID'], formatted_weight])
        current_row += len(local_weights_df) + 1
        current_row += 1

    # Write Local Weights Summary
    ws.cell(row=current_row, column=1, value="Bảng Tổng hợp Trọng số Cục bộ")
    current_row += 1
    ws.append(["Song ID"] + CRITERIA_NAMES)
    for i, row in local_weights_df.iterrows():
        formatted_row = []
        for x in row:
            if isinstance(x, (int, float)) and not np.isnan(x):
                formatted_row.append(f"{x:.4f}")
            else:
                formatted_row.append(str(x))
        ws.append([song_ids[i]] + formatted_row)
    current_row += len(song_ids) + 1

    # Add empty row
    current_row += 1

    # Write Final Results
    ws.cell(row=current_row, column=1, value="Kết quả Cuối cùng")
    current_row += 1
    ws.append(["Rank", "track_name", "track_artist", "Final Score", "Percentage"])
    for _, row in final_results_df.iterrows():
        final_score = row['Final Score']
        percentage = row.get('Percentage', 0)
        formatted_score = f"{final_score:.6f}" if isinstance(final_score, (int, float)) and not np.isnan(final_score) else str(final_score)
        formatted_percentage = f"{percentage:.2f}%" if isinstance(percentage, (int, float)) and not np.isnan(percentage) else str(percentage)
        ws.append([int(row['Rank']), row['track_name'], row['track_artist'], formatted_score, formatted_percentage])
    current_row += len(final_results_df) + 1

    # Add empty row
    current_row += 1

    # Write Charts
    ws.cell(row=current_row, column=1, value="Biểu đồ")
    current_row += 1
    try:
        for fig_name, fig in figures.items():
            img_buffer = BytesIO()
            if isinstance(fig, plt.Figure):
                fig.savefig(img_buffer, format='png', bbox_inches='tight')
                plt.close(fig)
            else:
                # Handle Plotly figure
                try:
                    pio.write_image(fig, file=img_buffer, format='png')
                except Exception as e:
                    st.error(f"Lỗi khi xuất biểu đồ '{fig_name}' sang PNG: {e}")
                    continue
            img_buffer.seek(0)
            img = Image(img_buffer)
            ws.add_image(img, f'A{current_row}')
            ws.cell(row=current_row, column=1, value=fig_name)
            # Adjust row height for images (assuming ~300 pixels height)
            ws.row_dimensions[current_row].height = 300 * 0.75  # Convert pixels to points (approx)
            current_row += 20  # Space for image (adjust based on image size)
    except Exception as e:
        st.error(f"Lỗi khi thêm biểu đồ vào Excel: {e}")

    # Save workbook to buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def export_to_pdf(criteria_matrix_df, criteria_weights_df, local_weights_dfs, final_results_df, figures):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, leading=20, alignment=1, spaceAfter=20, fontName='TimesNewRoman')
    heading_style = ParagraphStyle(name='Heading2', fontSize=12, leading=14, spaceAfter=12, fontName='TimesNewRoman')
    normal_style = styles['Normal']

    # Title
    elements.append(Paragraph("Kết quả Xếp hạng Bài hát bằng AHP", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    # User Inputs
    elements.append(Paragraph("Đầu vào người dùng", heading_style))
    data = [["Cặp Tiêu chí", "Giá trị"]]
    idx = 0
    for i in range(len(CRITERIA_IDS)):
        for j in range(i + 1, len(CRITERIA_IDS)):
            crit_left = CRITERIA_MAP[CRITERIA_IDS[i]]
            crit_right = CRITERIA_MAP[CRITERIA_IDS[j]]
            data.append([f"{crit_left} vs {crit_right}", f"{st.session_state.get('pairwise_inputs', [])[idx]:.2f}"])
            idx += 1
    table = Table(data, colWidths=[3.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Criteria Comparison Matrix
    elements.append(Paragraph("Ma trận So sánh Tiêu chí", heading_style))
    data = [[""] + CRITERIA_NAMES]
    for i, row in enumerate(criteria_matrix_df.values):
        formatted_row = [CRITERIA_NAMES[i]] + [f"{x:.3f}" if isinstance(x, (int, float)) and not np.isnan(x) else str(x) for x in row]
        data.append(formatted_row)
    table = Table(data, colWidths=[1.5*inch] + [1.0*inch]*len(CRITERIA_NAMES))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Criteria Weights
    elements.append(Paragraph("Trọng số Tiêu chí", heading_style))
    data = [["Tiêu chí", "Trọng số", "Rank"]]
    for _, row in criteria_weights_df.iterrows():
        weight = row['Trọng số (Weights)']
        formatted_weight = f"{weight:.4f}" if isinstance(weight, (int, float)) and not np.isnan(weight) else str(weight)
        data.append([row['Tiêu chí'], formatted_weight, str(int(row['Rank']))])
    table = Table(data, colWidths=[2.0*inch, 1.5*inch, 1.0*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Alternative Comparison Matrices and Local Weights
    for crit_name, local_weights_df in local_weights_dfs.items():
        elements.append(Paragraph(f"Ma trận So sánh Phương án - {crit_name}", heading_style))
        data = [[""] + local_weights_df['Song ID'].tolist()]
        for i, row in enumerate(local_weights_df['Matrix'].values):
            formatted_row = [local_weights_df['Song ID'][i]] + [f"{x:.3f}" if isinstance(x, (int, float)) and not np.isnan(x) else str(x) for x in row]
            data.append(formatted_row)
        table = Table(data, colWidths=[1.0*inch] + [0.8*inch]*len(local_weights_df))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.1 * inch))

        elements.append(Paragraph(f"Trọng số Phương án - {crit_name}", heading_style))
        data = [["Song ID", "Trọng số"]]
        for _, row in local_weights_df.iterrows():
            weight = row['Weight']
            formatted_weight = f"{weight:.4f}" if isinstance(weight, (int, float)) and not np.isnan(weight) else str(weight)
            data.append([row['Song ID'], formatted_weight])
        table = Table(data, colWidths=[2.0*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.2 * inch))

    # Local Weights Summary
    elements.append(Paragraph("Bảng Tổng hợp Trọng số Cục bộ", heading_style))
    data = [["Song ID"] + CRITERIA_NAMES]
    for i, row in local_weights_df.iterrows():
        formatted_row = [f"{x:.4f}" if isinstance(x, (int, float)) and not np.isnan(x) else str(x) for x in row]
        data.append([song_ids[i]] + formatted_row)
    table = Table(data, colWidths=[1.0*inch] + [1.0*inch]*len(CRITERIA_NAMES))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Final Results
    elements.append(Paragraph("Kết quả Cuối cùng", heading_style))
    data = [["Xếp hạng", "Tên bài hát", "Nghệ sĩ", "Điểm số", "Tỷ lệ %"]]
    for _, row in final_results_df.iterrows():
        final_score = row['Final Score']
        percentage = row.get('Percentage', 0)
        formatted_score = f"{final_score:.6f}" if isinstance(final_score, (int, float)) and not np.isnan(final_score) else str(final_score)
        formatted_percentage = f"{percentage:.2f}%" if isinstance(percentage, (int, float)) and not np.isnan(percentage) else str(percentage)
        data.append([str(int(row['Rank'])), row['track_name'], row['track_artist'], formatted_score, formatted_percentage])
    table = Table(data, colWidths=[0.8*inch, 2.0*inch, 1.5*inch, 1.2*inch, 1.0*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'TimesNewRoman'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Charts
    elements.append(Paragraph("Biểu đồ", heading_style))
    try:
        for fig_name, fig in figures.items():
            img_buffer = BytesIO()
            if isinstance(fig, plt.Figure):
                fig.savefig(img_buffer, format='png', bbox_inches='tight')
                plt.close(fig)
            else:
                # Handle Plotly figure
                try:
                    pio.write_image(fig, file=img_buffer, format='png', width=600, height=400)
                except Exception as e:
                    st.error(f"Lỗi khi xuất biểu đồ '{fig_name}' sang PNG: {e}")
                    continue
            img_buffer.seek(0)
            img = ReportLabImage(img_buffer, width=5.5*inch, height=3.5*inch)
            elements.append(Paragraph(fig_name, heading_style))
            elements.append(img)
            elements.append(Spacer(1, 0.2 * inch))
    except Exception as e:
        st.error(f"Lỗi khi thêm biểu đồ vào PDF: {e}")

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

st.set_page_config(layout="wide", page_title="AHP Đề xuất Bài hát")
st.title("🎵 Ứng dụng Đề xuất Bài hát bằng Phương pháp AHP")
st.write("""
Ứng dụng này sử dụng phương pháp Phân tích Thứ bậc (AHP) để giúp bạn xếp hạng các bài hát
dựa trên các tiêu chí và mức độ ưu tiên do bạn thiết lập.
""")

st.header(" Tải lên file CSV chứa dữ liệu bài hát")
uploaded_file = st.file_uploader("Chọn file CSV", type=["csv"])

if uploaded_file is not None:
    df_songs = load_data(uploaded_file)
    if df_songs is not None:
        st.sidebar.header("⚙️ Thiết lập AHP")
        st.sidebar.subheader("1. Mức độ quan trọng của Tiêu chí")
        st.sidebar.info("Hãy so sánh mức độ quan trọng của các cặp tiêu chí theo thang đo AHP (1-9 và nghịch đảo).")

        pairwise_inputs = []
        for i in range(len(CRITERIA_IDS)):
            for j in range(i + 1, len(CRITERIA_IDS)):
                crit_left = CRITERIA_MAP[CRITERIA_IDS[i]]
                crit_right = CRITERIA_MAP[CRITERIA_IDS[j]]
                selected_label = st.sidebar.selectbox(
                    f"'{crit_left}' quan trọng hơn '{crit_right}' bao nhiêu?",
                    options=AHP_LABELS,
                    index=AHP_VALUES.index(1),
                    key=f"compare_{i}_{j}"
                )
                selected_value = float(selected_label.split('(')[-1].replace(')', ''))
                pairwise_inputs.append(selected_value)
        st.session_state['pairwise_inputs'] = pairwise_inputs

        st.header("📊 Kết quả Tính toán AHP")
        st.subheader("Bước 1: Tính Trọng số Tiêu chí")
        criteria_comparison_matrix = create_criteria_matrix_from_inputs(pairwise_inputs, CRITERIA_IDS)
        criteria_matrix_df = pd.DataFrame(criteria_comparison_matrix, index=CRITERIA_NAMES, columns=CRITERIA_NAMES)

        with st.expander("Xem Ma trận So sánh Tiêu chí"):
            st.write("Ma trận dựa trên đánh giá của bạn:")
            st.dataframe(criteria_matrix_df.style.format("{:.3f}"))

        criteria_weights, crit_lambda_max, crit_ci, crit_cr = calculate_ahp_weights_consistency(criteria_comparison_matrix)
        st.write("Trọng số ưu tiên của các tiêu chí:")
        df_crit_weights = pd.DataFrame({
            'Tiêu chí': CRITERIA_NAMES,
            'Trọng số (Weights)': criteria_weights
        })
        df_crit_weights['Rank'] = df_crit_weights['Trọng số (Weights)'].rank(method='dense', ascending=False).astype(int)
        df_crit_weights = df_crit_weights.sort_values(by='Rank')
        st.dataframe(df_crit_weights.style.format({'Trọng số (Weights)': '{:.4f}'}).hide(axis="index"))

        st.subheader("Biểu đồ Trọng số Tiêu chí")
        fig_crit = px.bar(
            df_crit_weights,
            x='Tiêu chí',
            y='Trọng số (Weights)',
            text='Trọng số (Weights)',
            labels={'Tiêu chí': 'Tiêu chí', 'Trọng số (Weights)': 'Trọng số'},
            title='Trọng số của Các Tiêu chí (Plotly)',
            color='Trọng số (Weights)',
            color_continuous_scale='Blues'
        )
        fig_crit.update_traces(texttemplate='%{text:.4f}', textposition='auto')
        fig_crit.update_layout(
            xaxis_title="Tiêu chí",
            yaxis_title="Trọng số",
            xaxis_tickangle=45,
            showlegend=False,
            height=500
        )
        st.plotly_chart(fig_crit, use_container_width=True)
        st.session_state['figures'] = st.session_state.get('figures', {})
        st.session_state['figures']['Criteria Weights'] = fig_crit

        st.write(f"**Kiểm tra Nhất quán Tiêu chí:**")
        st.write(f"- Lambda Max (λmax): {crit_lambda_max:.4f}")
        st.write(f"- Chỉ số Nhất quán (CI): {crit_ci:.4f}")
        crit_ri = RI_LOOKUP.get(len(CRITERIA_IDS), "N/A")
        st.write(f"- Chỉ số Ngẫu nhiên (RI) cho n={len(CRITERIA_IDS)}: {crit_ri}")

        if isinstance(crit_ri, str) or crit_ri == 0:
            st.write(f"- Tỷ số Nhất quán (CR): {'Nhất quán' if np.isclose(crit_ci, 0) else 'Không xác định / Không nhất quán'}")
            is_consistent = np.isclose(crit_ci, 0)
        else:
            st.write(f"- Tỷ số Nhất quán (CR): {crit_cr:.4f}")
            is_consistent = crit_cr <= 0.10

        if is_consistent:
            st.success("-> Đánh giá tiêu chí là nhất quán (CR <= 0.10)")
        else:
            st.warning("-> Đánh giá tiêu chí KHÔNG nhất quán (CR > 0.10). Vui lòng xem xét lại các so sánh cặp.")
            st.stop()

        if is_consistent:
            if st.button("Lưu Trọng Số Tiêu Chí và Tiếp Tục"):
                st.session_state['weights_saved'] = True
                st.success("Trọng số tiêu chí đã được lưu.")
                st.session_state['criteria_weights'] = criteria_weights
                st.session_state['criteria_matrix_df'] = criteria_matrix_df
                st.session_state['criteria_weights_df'] = df_crit_weights

        if 'weights_saved' in st.session_state and st.session_state['weights_saved']:
            st.sidebar.subheader("2. Lựa chọn Phương án (Bài hát)")
            max_songs = len(df_songs)
            num_songs_to_compare = st.sidebar.number_input(
                f"Chọn số lượng bài hát hàng đầu để so sánh (tối đa {max_songs})",
                min_value=2,
                max_value=max_songs,
                value=min(4, max_songs),
                step=1,
                key="num_songs"
            )

            selected_songs = df_songs.head(num_songs_to_compare).copy()
            global song_ids
            song_ids = [f"BH{i+1}" for i in range(num_songs_to_compare)]
            selected_songs['song_id'] = song_ids

            st.sidebar.write("---")
            st.sidebar.subheader("Các Bài hát được chọn:")
            st.sidebar.dataframe(selected_songs[['song_id', 'track_name', 'track_artist']], hide_index=True)

            st.write("---")
            st.subheader(f"Bước 2: Đánh giá {num_songs_to_compare} Bài hát theo Từng Tiêu chí")

            local_weights_data = {}
            local_weights_dfs = {}
            consistency_results_alt = {}
            criteria_col_map = {
                "phobien": "track_popularity",
                "nhipdo": "tempo",
                "nangluong": "energy",
                "nhay": "danceability",
                "camxuc": "valence"
            }
            all_alternatives_consistent = True

            for crit_id, crit_name in CRITERIA_MAP.items():
                col_name = criteria_col_map[crit_id]
                st.markdown(f"**Theo tiêu chí: {crit_name}** (sử dụng cột `{col_name}`)")
                song_values = selected_songs[col_name].values
                alt_comparison_matrix = create_alternative_comparison_matrix(song_values, crit_id)
                local_weights, alt_lambda_max, alt_ci, alt_cr = calculate_ahp_weights_consistency(alt_comparison_matrix)
                local_weights_data[crit_name] = local_weights
                consistency_results_alt[crit_name] = {'lambda': alt_lambda_max, 'ci': alt_ci, 'cr': alt_cr}
                alt_df = pd.DataFrame({
                    'Song ID': song_ids,
                    'Weight': local_weights,
                    'Matrix': list(alt_comparison_matrix)
                })
                local_weights_dfs[crit_name] = alt_df
                with st.expander(f"Xem chi tiết tính toán cho '{crit_name}'"):
                    st.write("Ma trận so sánh bài hát (dựa trên tỷ lệ giá trị):")
                    st.dataframe(pd.DataFrame(alt_comparison_matrix, index=song_ids, columns=song_ids).style.format("{:.3f}"))
                    st.write("Trọng số cục bộ (Local Weights - CW):")
                    st.dataframe(alt_df[['Song ID', 'Weight']].style.format({'Weight': '{:.4f}'}).hide(axis='index'))
                    st.write(f"Kiểm tra nhất quán (CR={alt_cr:.4f} - Thường bằng 0 do tính toán từ tỷ lệ).")
                    if not np.isclose(alt_cr, 0.0, atol=1e-5) and alt_cr > 0.10:
                        all_alternatives_consistent = False

            global local_weights_df
            local_weights_df = pd.DataFrame(local_weights_data, index=song_ids)
            st.session_state['local_weights_dfs'] = local_weights_dfs
            st.session_state['local_weights_df'] = local_weights_df

            st.write("---")
            st.subheader("Bảng Tổng hợp Trọng số Cục bộ (CW) của Bài hát")
            st.dataframe(local_weights_df.style.format("{:.4f}"))
            if not all_alternatives_consistent:
                st.warning("Cảnh báo: Có ít nhất một ma trận so sánh phương án không nhất quán (CR > 0.10).")

            st.write("---")
            st.subheader("Bước 3: Kết quả Xếp hạng Cuối cùng")

            try:
                local_weights_matrix = local_weights_df[CRITERIA_NAMES].values
                criteria_weights_vector = st.session_state['criteria_weights']
                if local_weights_matrix.shape[1] == criteria_weights_vector.shape[0]:
                    final_scores = local_weights_matrix @ criteria_weights_vector
                    final_results_df = pd.DataFrame({
                        'Song ID': song_ids,
                        'Final Score': final_scores
                    })
                    final_results_df = final_results_df.merge(
                        selected_songs[['song_id', 'track_name', 'track_artist']],
                        left_on='Song ID',
                        right_on='song_id',
                        how='left'
                    ).drop(columns=['song_id'])
                    final_results_df['Rank'] = final_results_df['Final Score'].rank(method='dense', ascending=False).astype(int)
                    final_results_df = final_results_df.sort_values(by='Rank')
                    total_score = final_results_df['Final Score'].sum()
                    final_results_df['Percentage'] = (final_results_df['Final Score'] / total_score * 100).round(2)

                    st.dataframe(
                        final_results_df[['Rank', 'track_name', 'track_artist', 'Final Score', 'Percentage']].style.format({'Final Score': '{:.6f}', 'Percentage': '{:.2f}%'}),
                        hide_index=True,
                        use_container_width=True
                    )

                    best_song = final_results_df.iloc[0]
                    st.success(f"🏆 Bài hát được đề xuất hàng đầu: **{best_song['track_name']}** của {best_song['track_artist']} (Điểm: {best_song['Final Score']:.6f})")

                    st.subheader("Biểu đồ Điểm số Bài hát (Plotly)")
                    final_results_df['Label'] = final_results_df['track_name'] + ' (' + final_results_df['track_artist'] + ')'
                    fig = px.bar(
                        final_results_df,
                        x='Label',
                        y='Final Score',
                        text='Final Score',
                        labels={'Label': 'Bài hát', 'Final Score': 'Điểm số'},
                        title='Điểm số Cuối cùng của Các Bài hát (Plotly)',
                        color='Final Score',
                        color_continuous_scale='Viridis'
                    )
                    fig.update_traces(texttemplate='%{text:.6f}', textposition='auto')
                    fig.update_layout(
                        xaxis_title="Bài hát",
                        yaxis_title="Điểm số Cuối cùng",
                        xaxis_tickangle=45,
                        showlegend=False,
                        height=500
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    st.session_state['figures']['Final Scores Bar'] = fig

                    st.subheader("Biểu đồ Tỷ lệ Phần trăm Điểm số Bài hát")
                    fig_pie = px.pie(
                        final_results_df,
                        names='Label',
                        values='Percentage',
                        title='Tỷ lệ Phần trăm Điểm số của Các Bài hát',
                        color_discrete_sequence=px.colors.sequential.RdBu
                    )
                    fig_pie.update_traces(textinfo='percent+label', pull=[0.1 if i == 0 else 0 for i in range(len(final_results_df))])
                    fig_pie.update_layout(
                        showlegend=True,
                        height=500
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)
                    st.session_state['figures']['Final Scores Pie'] = fig_pie

                    st.subheader("Biểu đồ Điểm số Bài hát (Matplotlib)")
                    fig, ax = plt.subplots(figsize=(10, 6))
                    bars = ax.bar(final_results_df['Label'], final_results_df['Final Score'], color='skyblue', edgecolor='black')
                    for bar in bars:
                        yval = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2, yval, f'{yval:.6f}', 
                                ha='center', va='bottom', fontsize=10)
                    ax.set_xlabel('Bài hát', fontsize=12)
                    ax.set_ylabel('Điểm số Cuối cùng', fontsize=12)
                    ax.set_title('Điểm số Cuối cùng của Các Bài hát (Matplotlib)', fontsize=14)
                    ax.tick_params(axis='x', rotation=45, labelsize=10)
                    ax.grid(True, axis='y', linestyle='--', alpha=0.7)
                    plt.tight_layout()
                    st.pyplot(fig)
                    st.session_state['figures']['Final Scores Matplotlib'] = fig

                    st.session_state['final_results_df'] = final_results_df

                    st.subheader("Lưu Kết quả vào Cơ sở Dữ liệu")
                    if st.button("Lưu Dữ liệu"):
                        connection = init_mysql_connection()
                        if connection:
                            create_results_table(connection)
                            save_results_to_mysql(connection, final_results_df)
                            connection.close()

                    st.subheader("Xuất Kết quả ra Excel")
                    if st.button("Xuất ra Excel"):
                        excel_buffer = export_to_excel(
                            st.session_state['criteria_matrix_df'],
                            st.session_state['criteria_weights_df'],
                            st.session_state['local_weights_dfs'],
                            st.session_state['final_results_df'],
                            st.session_state['figures']
                        )
                        st.download_button(
                            label="Tải file Excel",
                            data=excel_buffer,
                            file_name="AHP_Song_Ranking_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    st.subheader("Xuất Kết quả ra PDF")
                    if st.button("Xuất ra PDF"):
                        pdf_buffer = export_to_pdf(
                            st.session_state['criteria_matrix_df'],
                            st.session_state['criteria_weights_df'],
                            st.session_state['local_weights_dfs'],
                            st.session_state['final_results_df'],
                            st.session_state['figures']
                        )
                        st.download_button(
                            label="Tải file PDF",
                            data=pdf_buffer,
                            file_name="AHP_Song_Ranking_Results.pdf",
                            mime="application/pdf"
                        )

                    # Hiển thị lịch sử dữ liệu
                    st.subheader("Lịch sử Kết quả")
                    connection = init_mysql_connection()
                    if connection:
                        history_df = load_history_from_mysql(connection)
                        if not history_df.empty:
                            st.dataframe(
                                history_df.style.format({'Final Score': '{:.6f}', 'saved_at': '{:%Y-%m-%d %H:%M:%S}'}),
                                hide_index=True,
                                use_container_width=True
                            )
                        else:
                            st.write("Chưa có dữ liệu lịch sử.")
                        connection.close()
                else:
                    st.error(f"Lỗi kích thước không khớp để nhân ma trận:")
                    st.error(f"- Kích thước ma trận trọng số cục bộ: {local_weights_matrix.shape}")
                    st.error(f"- Kích thước vector trọng số tiêu chí: {criteria_weights_vector.shape}")
                    st.error("Vui lòng kiểm tra lại thứ tự tiêu chí và tính toán.")

            except KeyError as e:
                st.error(f"Lỗi: Không tìm thấy cột tiêu chí '{e}' trong bảng trọng số cục bộ.")
                st.error("Đảm bảo tên tiêu chí nhất quán trong suốt quá trình tính toán.")
            except Exception as e:
                st.error(f"Đã xảy ra lỗi không mong muốn trong quá trình tổng hợp: {e}")

        else:
            st.info("Vui lòng lưu trọng số tiêu chí trước khi tiếp tục.")

    else:
        st.warning("Không thể tải dữ liệu bài hát. Vui lòng kiểm tra file CSV.")

st.markdown("---")
st.caption("Ứng dụng Streamlit tạo bởi Mô hình Ngôn ngữ Lớn dựa trên notebook AHP.")